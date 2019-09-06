from intuitlib.client import AuthClient
from quickbooks import QuickBooks
from quickbooks.objects import *
from quickbooks.batch import *
import os
import tkinter as tk
from tkinter import filedialog
import datetime
import openpyxl
import qbapi
'''
auth_client = AuthClient(
        client_id='ABD2a9rkkwbLIK1xc4yG1OpmxlbFpTjRtVQJONAcdKThNnHC4a',
        client_secret='Njox3cMmbFLovPdpIcJ5XEyP5yURFaxEoQi6pmaG',
        environment='sandbox',
        redirect_uri='https://Quickbooks-API--loganmr.repl.co/sampleappoauth2/authCodeHandler',
    )

client = QuickBooks(
         auth_client=auth_client,
         refresh_token='AB11576357896MjOYQSwRMUMDFb8wOl9sSOLrfNtYGBsYFoUXA',
         company_id='4611809164062348683',
    )
'''
root = tk.Tk()
root.withdraw()
FNameIn = filedialog.askopenfilename(title = "Open upload sheet", defaultextension = ".xlsx", filetypes = [("Excel workbook (.xlsx)",".xlsx")])
root.destroy()

wb = openpyxl.load_workbook(FNameIn)

try:
    wssr = wb["Sales Receipts"]
except:
    input("Select a valid excel file\nPress Enter to restart")
    os.execl("QBUpload.py")

colvars = []
for k in wssr[1]:
	colvars.append(k.value.replace(" ",""))

RESREC = []
for k in range(wssr.max_row-1):
    RESREC.append({})
    for j in range(len(colvars)):
        RESREC[k][colvars[j]] = wssr[openpyxl.utils.cell.get_column_letter(j+1)+str(k+2)].value
        
neceInfoItems = ['Rental Income','Cleaning Fare','Colorado Springs City Sales Tax','Colorado Springs District State Sales Tax','HomeAway Commission','Guest Booking Fees','Booking.com Commissions']
neceInfoVars = ['NETINCOME','CLEANINGFARE','CITYTAX','LOCALTAX','HOMEAWAYCOMMISSION','GUESTBOOKINGFEE','BOOKINGDOTCOMCOMMISSION']

SRList = []
for k in RESREC:
    linelist = []
    for j in range(7):
        linelist.append({'account':neceInfoItems[j], 'class':k['SOURCE1'], 'amount':k[neceInfoVars[j]], 'description':k['DESCRIPTION1']})
    '''SRtempObj = NewSR('GUESTSNAME','CHECKIN','LISTINGSNICKNAME',linelist,False,client)
    SRList.append(SRtempObj)'''
    SRList.append(linelist)

'''
results = batch_create(SRList, qb=client)
'''
    
    
    
    
    
    
    
    
    
    
    