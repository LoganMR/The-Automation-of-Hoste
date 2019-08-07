import requests
import datetime
import openpyxl
import sys
import tkinter as tk
import calendar
import json
import gucal #This is a local file with useful functions, not an actual package.
from tkinter import filedialog
from openpyxl import Workbook
from requests.auth import HTTPBasicAuth
from tkcalendar import DateEntry

def datetostringlist(k):
    return(k.isoformat())

def find(ar,key,res,dic):
	for k in ar:
		if k[key] == res:
			return(k)
	return(dic)

def afterFees(source,amount):
    k = 0
    l = 0
    if source in ['HomeAway', 'Booking.com', 'website', 'Web Direct']:
        k = round(amount*0.029+0.30,2)
    if source not in ['Airbnb','tripadvisor.com']:
        l = round(amount*0.003,2)
    return(round(amount - k - l,2))

def bookingFee(amount,date):
    if date > "2019-07-23":
        return(round(0.08*amount/1.08,2))
    else:
        return(0)

root = tk.Tk()
root.withdraw()

FNameOut = filedialog.asksaveasfilename(title = "Output Filename", defaultextension = ".xlsx", filetypes = [("Excel workbook (.xlsx)",".xlsx")])
root.destroy()

dates = gucal.dateselector()
dates = list(map(datetostringlist,dates))

RESERVATIONS = gucal.resdats(*dates, con='confirmed')

FormRess = [0]*len(RESERVATIONS)

for k in range(len(RESERVATIONS)):
    FormRess[k] = {}

for k in range(len(RESERVATIONS)):
    FormRess[k]['GUESTSNAME'] = RESERVATIONS[k]['guest']['fullName']

for k in range(len(RESERVATIONS)):
    FormRess[k]['CHECKIN'] = datetime.datetime.strptime(RESERVATIONS[k]['checkIn'][0:10], '%Y-%m-%d').date().strftime('%m/%d/%Y')

for k in range(len(RESERVATIONS)):
    FormRess[k]['CHECKOUT'] = datetime.datetime.strptime(RESERVATIONS[k]['checkOut'][0:10], '%Y-%m-%d').date().strftime('%m/%d/%Y')

for k in range(len(RESERVATIONS)):
    FormRess[k]['LISTINGSNICKNAME'] = RESERVATIONS[k]['listing']['nickname']

for k in range(len(RESERVATIONS)):
    FormRess[k]['DESCRIPTION'] = "(Nights: " + str(RESERVATIONS[k]['nightsCount']) + ") Dates of Stay " + FormRess[k]['CHECKIN'] + " - " + FormRess[k]['CHECKOUT']

for k in range(len(RESERVATIONS)):
    FormRess[k]['SOURCE'] = RESERVATIONS[k]['source']

for k in range(len(RESERVATIONS)):
    FormRess[k]['CONFIRMATIONDATE'] = RESERVATIONS[k]['confirmedAt'][0:10]

for k in range(len(RESERVATIONS)):
    FormRess[k]['CLEANINGFARE'] = find(RESERVATIONS[k]['money']['invoiceItems'],'title','Cleaning fee',{'amount':0})['amount']

for k in range(len(RESERVATIONS)):
    FormRess[k]['NETINCOME'] = RESERVATIONS[k]['money']['netIncome']
    
for k in range(len(RESERVATIONS)):
    FormRess[k]['ACCOMMODATIONFARE'] = RESERVATIONS[k]['money']['fareAccommodation']

for k in range(len(RESERVATIONS)):
    FormRess[k]['SUBTOTALPRICE'] = RESERVATIONS[k]['money']['subTotalPrice']

for k in range(len(RESERVATIONS)):
    FormRess[k]['YOURCOMMISSION'] = RESERVATIONS[k]['money']['commission']

for k in range(len(RESERVATIONS)):
    FormRess[k]['TOTALPAYOUT'] = RESERVATIONS[k]['money']['hostPayout']

for k in range(len(RESERVATIONS)):
    FormRess[k]['COMMISSIONRATE'] = float(RESERVATIONS[k]['money']['commissionFormula'][11:15])

for k in range(len(RESERVATIONS)):
    FormRess[k]['CITYTAX'] = find(RESERVATIONS[k]['money']['invoiceItems'],'title','CITY_TAX',{'amount' : 0})['amount']    

for k in range(len(RESERVATIONS)):
	j = find(RESERVATIONS[k]['money']['invoiceItems'],'title','LOCAL_TAX',{'amount' : 'none'})['amount']
	if j == 'none':
		j = find(RESERVATIONS[k]['money']['invoiceItems'],'title','State Tax',{'amount' : 0})['amount']
	FormRess[k]['LOCALTAX'] = j

for k in range(len(RESERVATIONS)):
    FormRess[k]['BOOKINGFEE'] = bookingFee(FormRess[k]['ACCOMMODATIONFARE'],FormRess[k]['CONFIRMATIONDATE'])

for k in range(len(RESERVATIONS)):
    if FormRess[k]['SOURCE'] == 'HomeAway':
        FormRess[k]['HOMEAWAY'] = round(FormRess[k]['SUBTOTALPRICE']*0.05,2)
    else:
        FormRess[k]['HOMEAWAY'] = 0

for k in range(len(RESERVATIONS)):
    FormRess[k]['NEWNETINCOME'] = round(afterFees(FormRess[k]['SOURCE'],FormRess[k]['TOTALPAYOUT'])-(FormRess[k]['CLEANINGFARE']+FormRess[k]['CITYTAX']+FormRess[k]['LOCALTAX']+FormRess[k]['HOMEAWAY']+FormRess[k]['BOOKINGFEE']),2)

for k in range(len(RESERVATIONS)):
    FormRess[k]['NEWCOMMISSION'] = round(FormRess[k]['NEWNETINCOME']*FormRess[k]['COMMISSIONRATE'],2)

for k in range(len(RESERVATIONS)):
    FormRess[k]['HOSTENAME'] = 'Hoste, LLC'

for k in FormRess:
	if '(Own' in k['LISTINGSNICKNAME']:
		FormRess.remove(k)

wb2 = Workbook()
ws1 = wb2.create_sheet("Sales Receipts")
ws2 = wb2.create_sheet("Journal Entries")
ws1.title = "Sales Receipts"
ws2.title = "Journal Entries"
wb2.remove(wb2['Sheet'])

ws1['A1'] = 'CHECK IN'
ws1['B1'] = 'CHECK OUT' 
ws1['C1'] = 'LISTINGS NICKNAME' 
ws1['D1'] = 'GUESTS NAME' 
ws1['E1'] = 'CLEANINGFARE'
ws1['F1'] = 'NET INCOME'
ws1['G1'] = 'CITY TAX'
ws1['H1'] = 'LOCAL TAX'
ws1['I1'] = 'SOURCE 1'
ws1['J1'] = 'SOURCE 2'
ws1['K1'] = 'SOURCE 3'
ws1['L1'] = 'SOURCE 4'
ws1['M1'] = 'SOURCE 5'
ws1['N1'] = 'SOURCE 6'
ws1['O1'] = 'DESCRIPTION 1'
ws1['P1'] = 'DESCRIPTION 2'
ws1['Q1'] = 'DESCRIPTION 3'
ws1['R1'] = 'DESCRIPTION 4'
ws1['S1'] = 'DESCRIPTION 5'
ws1['T1'] = 'DESCRIPTION 6'
ws1['U1'] = 'HOMEAWAY COMMISSION'
ws1['V1'] = 'GUEST BOOKING FEE'

AR1 = ['CHECKIN','CHECKOUT','LISTINGSNICKNAME','GUESTSNAME','CLEANINGFARE','NEWNETINCOME','CITYTAX','LOCALTAX','SOURCE','SOURCE','SOURCE','SOURCE','SOURCE','SOURCE','DESCRIPTION','DESCRIPTION','DESCRIPTION','DESCRIPTION','DESCRIPTION','DESCRIPTION','HOMEAWAY','BOOKINGFEE']
for c in range(len(AR1)):
    for k in range(len(FormRess)):
        ws1[openpyxl.utils.cell.get_column_letter(c+1)+str(k+2)] = eval("FormRess[k]['" + AR1[c] + "']")

ws2['A1'] = 'CHECK IN'
ws2['B1'] = 'CHECK OUT'
ws2['C1'] = 'LISTINGS NICKNAME' 
ws2['D1'] = 'YOUR COMMISSION 1' 
ws2['E1'] = 'YOUR COMMISSION 2'
ws2['F1'] = 'DESCRIPTION 1' 
ws2['G1'] = 'DESCRIPTION 2' 
ws2['H1'] = 'HOSTENAME 1' 
ws2['I1'] = 'HOSTENAME 2' 
ws2['J1'] = 'PROPERTY NICKNAME 1' 
ws2['K1'] = 'PROPERTY NICKNAME 2' 
ws2['L1'] = 'SOURCE 1'
ws2['M1'] = 'SOURCE 2'

AR2 = ['CHECKIN','CHECKOUT','LISTINGSNICKNAME','NEWCOMMISSION','NEWCOMMISSION','DESCRIPTION','DESCRIPTION','HOSTENAME','HOSTENAME','LISTINGSNICKNAME','LISTINGSNICKNAME','SOURCE','SOURCE']
for c in range(len(AR2)):
    for k in range(len(FormRess)):
        ws2[openpyxl.utils.cell.get_column_letter(c+1)+str(k+2)] = eval("FormRess[k]['" + AR2[c] + "']")

wb2.save(FNameOut)
