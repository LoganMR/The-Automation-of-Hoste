# PLEASE READ:
# This code only works if the columns to be included on Guesty for the .csv are SELECTED in the following order: 
# CHECK IN 
# CHECK OUT
# CONFIRMATION CODE
# LISTING'S NICKNAME
# GUEST'S NAME
# GUEST'S EMAIL
# CLEANING FARE
# COMMISSION
# NET INCOME
# CITY TAX
# LOCAL TAX
# BALANCE DUE
# SOURCE
# ACCOMODATION FARE
# TOTAL PAYOUT

import requests
import datetime
import openpyxl
from requests.auth import HTTPBasicAuth

FName = input('File to edit (without .xlsx)')
FNameOut = input('Name of output file (without .xlsx)')

# -------------- Begin: Open workbook, worksheet, and make a new sheet

wb = openpyxl.load_workbook(FName+'.xlsx')
ws=wb.active
ws.title = 'Sales Receipts'
ws2 = wb.create_sheet('Journal Entry')
ws2.title = 'Journal Entry'

# -------------- End: Open workbook, worksheet, and make a new sheet

# -------------- Begin: Open HomeAway's Excel Sheet to get commissions

today = datetime.datetime.now()
month = today.strftime("%B")

wb = openpyxl.load_workbook(FName+'.xlsx')
ws=wb.active
ws.title = 'Sales Receipts'
ws2 = wb.create_sheet('Journal Entry')
ws2.title = 'Journal Entry'

# -------------- End: Open HomeAway's Excel Sheet to get commissions

# -------------- Begin: Remove nonetypes

for row in ws:
    for cell in row:
        if str(type(cell.value)) == "<class 'NoneType'>":
            cell.value = 0

# -------------- End: Remove nonetypes

# -------------- Begin: Store all columns as arrays

CHECKIN = []
for cell in ws['A']:
    CHECKIN.append(cell.value)

CHECKOUT = []
for cell in ws['B']:
    CHECKOUT.append(cell.value)

CONFIRMATIONCODE = []
for cell in ws['C']:
    CONFIRMATIONCODE.append(cell.value)

LISTINGSNICKNAME = []
for cell in ws['D']:
    LISTINGSNICKNAME.append(cell.value)

GUESTSNAME = []
for cell in ws['E']:
    GUESTSNAME.append(cell.value)

GUESTSEMAIL = []
for cell in ws['F']:
    GUESTSEMAIL.append(cell.value)
        
CLEANINGFARE = []
for cell in ws['G']:
    CLEANINGFARE.append(cell.value)

CITYTAX = []
for cell in ws['J']:
    CITYTAX.append(cell.value)
    
LOCALTAX = []
for cell in ws['K']:
    LOCALTAX.append(cell.value)

BALANCEDUE = []
for cell in ws['L']:
    BALANCEDUE.append(cell.value)

SOURCE = []
for cell in ws['M']:
    SOURCE.append(cell.value)

ACCOMODATION = []
for cell in ws['N']:
    ACCOMODATION.append(cell.value)

TOTALPAYOUT = []
for cell in ws['O']:
    TOTALPAYOUT.append(cell.value)

SUBTOTALPRICE = []
for cell in ws['P']:
    SUBTOTALPRICE.append(cell.value)

# -------------- End: Store all columns as arrays

# -------------- Begin: Calculate the number of nights a guest stays

NIGHTS = ['NIGHTS']
for k in range(1,ws.max_row):
	NIGHTS.append((CHECKOUT[k]-CHECKIN[k]).days)

# -------------- End: Calculate the number of nights a guest stays

# -------------- Begin: Create Descriptions

DESCRIPTION = ['DESCRIPTION']
for k in range(1,ws.max_row):
    DESCRIPTION.append('(Nights: '+str(NIGHTS[k]) + ') Dates of Stay: ' + ws['A'+str(k+1)].value.strftime('%m/%d/%y') + ' - ' + ws['B'+str(k+1)].value.strftime('%m/%d/%y'))

# -------------- End: Create Descriptions
# -----------------------------------------------------------------------
# -------------- Begin: Add column titles

ws.insert_cols(14,7)
ws['M1'] = "SOURCE 1"
ws['N1'] = "SOURCE 2"
ws['O1'] = "SOURCE 3"
ws['P1'] = "SOURCE 4"
ws['Q1'] = "DESCRIPTION 1"
ws['R1'] = "DESCRIPTION 2"
ws['S1'] = "DESCRIPTION 3"
ws['T1'] = "DESCRIPTION 4"
ws['X1'] = "HOMEAWAY COMMISSION"

# -------------- End: Add column titles 

# -------------- Begin: Copy sources

for k in range(1,ws.max_row):
    ws['N'+str(k+1)].value = SOURCE[k]
    ws['O'+str(k+1)].value = SOURCE[k]
    ws['P'+str(k+1)].value = SOURCE[k]
        
# -------------- End: Copy sources

# -------------- Begin: Create Descriptions

for k in range(1,ws.max_row):
    ws['Q'+str(k+1)] = DESCRIPTION[k]
    ws['R'+str(k+1)] = DESCRIPTION[k]
    ws['S'+str(k+1)] = DESCRIPTION[k]
    ws['T'+str(k+1)] = DESCRIPTION[k]

# -------------- End: Create Descriptions
# ----------------------------------------------------------------------
flist_Path = "C:/Users/Logan Robinson/Desktop/GuestyScripts/FeeList"
flist_F = open(flist_Path)
fT = flist_F.read()

EXE = []
exelis = []
ein = fT.find('Exemptions')
ein = fT.find("\n",ein+1)
exelis.append(ein)
nlin = fT.find("\n",ein+1)
while ein>0:
    ein = fT.find(", ",ein+1, nlin)
    exelis.append(ein)
    
for k in range(0,len(exelis)-2):
    EXE.append(fT[exelis[k]+2:exelis[k+1]])

THR = []
thrlis = []
tin = fT.find('Sources')
tin = fT.find("\n",tin+1)
thrlis.append(tin)
nlin = fT.find("\n",tin+1)
while tin>0:
    tin = fT.find(", ",tin+1,nlin)
    thrlis.append(tin)
    
for k in range(0,len(thrlis)-2):
    THR.append(fT[thrlis[k]+2:thrlis[k+1]])

finlis = []
fin = fT.find("BEGIN_LIST")
while fin>0:
    fin = fT.find("\n",fin+1)
    finlis.append(fin)

flist_Dic = {}
for k in range(0,len(finlis)-2):
    Divr = fT.find("__",finlis[k],finlis[k+1])
    flist_Dic[fT[finlis[k]+1:Divr]] = fT[Divr+2:finlis[k+1]]

def throughStripe(src):
    if src in THR:
        return(True)

def fees(src,N):
    Form = flist_Dic[src]
    Divr = Form.find("#")
    h1 = Form[0:Divr]
    h2 = Form[Divr+1:]
    h = h1 + str(N) + h2
    h = eval(h)
    return(round(h,2))

def afterFees(src,N):
    k = 0
    l = 0
    if src in THR:
        k = fees('Stripe',N)
    if src not in EXE:
        l = fees('Guesty',N)
    return(round(N - k - l,2))
# --------------- End: Create function to calculate fees
# ----------------------------------------------------------------------
# --------------- Begin: Create a dictionary of property names : rates

plist_Path = "C:/Users/Logan Robinson/Desktop/GuestyScripts/PropertyList.txt"
plist_F = open(plist_Path)
T = plist_F.read()

linlis = []
lin = T.find("BEGIN_LIST")
while lin>0:
    lin = T.find("\n",lin+1)
    linlis.append(lin)

plist_Dic = {}
for k in range(0,len(linlis)-2):
    Divr = T.find("__",linlis[k],linlis[k+1])
    plist_Dic[T[linlis[k]+1:Divr]] = float(T[Divr+2:linlis[k+1]])

# --------------- End: Create a dictionary of property names : rates
# -----------------------------------------------------------------------
# -------------- Begin: Create HomeAway fee

HOMEAWAY = ['HOMEAWAY COMMISSION']
for k in range(1,ws.max_row):
    if SOURCE[k] == 'HomeAway':
        HOMEAWAY.append(fees('HomeAway',SUBTOTALPRICE[k]))
    else:
        HOMEAWAY.append(0)

# -------------- End: Create HomeAway fee

# --------------- Begin: Calculate Net Income by remainders after taxes and cleaning fare

NETINCOME = ['NET INCOME']
for k in range(1,ws.max_row):
    NETINCOME.append(round(afterFees(SOURCE[k],TOTALPAYOUT[k])-(CLEANINGFARE[k]+CITYTAX[k]+LOCALTAX[k]+HOMEAWAY[k]),2))

# --------------- End: Calculate Net Income by remainders after taxes and cleaning fare
# --------------- Begin: Calculate Net Income by remainders after taxes and cleaning fare

COMMISSION = ['YOUR COMMISSION']
for k in range(1,ws.max_row):
    COMMISSION.append(round(NETINCOME[k]*plist_Dic[LISTINGSNICKNAME[k]],2))

# --------------- End: Calculate Net Income by remainders after taxes and cleaning fare
# --------------- Begin: Populate NET INCOME, COMMISSION, and HOMEAWAY COMMISSION

for cell in ws['H']:
    cell.value = COMMISSION[cell.row-1]
    
for cell in ws['I']:
    cell.value = NETINCOME[cell.row-1]

for cell in ws['X']:
    cell.value = HOMEAWAY[cell.row-1]

# --------------- End: Populate NET INCOME, COMMISSION, and HOMEAWAY COMMISSION
# -----------------------------------------------------------------------
# --------------- Begin: Duplicate worksheet and adjust columns

ws2['A1'] = 'CHECK IN'
ws2['B1'] = 'CHECK OUT'

for row in ws: # Date cells copy as strings so we need to do check-in and check-out columns seperately
    for cell in row:
        if cell.column > 2:
            ws2[openpyxl.utils.cell.get_column_letter(cell.column)+str(cell.row)] = cell.value

for cell in ws['A']:
    if cell.row > 1:
        ws2[openpyxl.utils.cell.get_column_letter(cell.column)+str(cell.row)] = datetime.datetime.strptime(str(cell.value)[0:10],"%Y-%m-%d").date()
for cell in ws['B']:
    if cell.row > 1:
        ws2[openpyxl.utils.cell.get_column_letter(cell.column)+str(cell.row)] = datetime.datetime.strptime(str(cell.value)[0:10],"%Y-%m-%d").date()
        
ws2.insert_cols(5)
ws2.insert_cols(10)

for cell in ws2['E']:
    cell.value = ws2['D'+str(cell.row)].value
    
for cell in ws2['J']:
    cell.value = ws2['I'+str(cell.row)].value

ws2['D1'] = "LISTING'S NICKNAME 1"
ws2['E1'] = "LISTING'S NICKNAME 2"

ws2['I1'] = "YOUR COMMISSION 1"
ws2['J1'] = "YOUR COMMISSION 2"

ws2.delete_cols(21,2)
ws2.delete_cols(17,2)

ws2['U1'] = "HOSTENAME 1"
ws2['V1'] = "HOSTENAME 2"

for k in range(2,ws.max_row+1):
    ws2['U'+str(k)].value = 'Hoste, LLC'
    ws2['V'+str(k)].value = 'Hoste, LLC'
# --------------- End: Duplicate worksheet and adjust columns

wb.save(FNameOut+'.xlsx')